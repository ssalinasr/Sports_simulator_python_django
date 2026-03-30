from core_scripts.leagues import league_group
from core_scripts.tournaments import tournament_group
from core_scripts.leagues import league_tools
import random
import itertools
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse
from appolympics.models import Nationalteams, Olympicplayers, Teammatchesregister, Teamsports, Playertournamentsports, Teamtournamentregister, Playertournamentregister

class FullTournament():

    def __init__(self, teams, divisions, ranks, sport, num_groups, match_class, year, saveres):
        self.teams = teams
        self.divisions = divisions
        self.ranks = ranks
        self.sport = sport
        self.match_class = match_class
        self.year = year
        self.saveres = saveres
        self.prev_stages = []
        self.prev_stages_tables = []
        self.prev_stages_matches = []
        self.qualified_teams = []
        self.group_finals = []
        self.group_finals_tables = []
        self.group_finals_matches = []
        self.world_groups = []
        self.world_groups_tables = []
        self.world_groups_matches = []
        self.final = []
        self.final_table = {}
        self.final_matches = []
        self.element_names = []
        self.num_groups = num_groups
        self.general_tables = []
        self.general_matches = []

    def subdivide_groups(self, num_groups, teams):
        random.shuffle(teams)
        base_length = len(teams) // num_groups
        remainder = len(teams) % num_groups

        sorted_teams = []
        index = 0

        for i in range(num_groups):
            length = base_length + 1 if i < remainder else base_length
            sorted_teams.append(teams[index:index+length])
            index += length
        
        return sorted_teams

    def simulate_tournament(self):
        if isinstance(self.divisions, list): # Simula torneos de deportes entre paises #
            region_sublist = ['Africa', 'Asia_Oceania', 'America', 'Europa']
            #Genera los grupos de la fase previa#
            for div in range(len(self.divisions)):
                self.prev_stages.append(self.subdivide_groups(self.num_groups, self.teams[div]))
            counter = 0
            #Simula los grupos de la fase previa#
            for stage_group in self.prev_stages:
                index = 1
                temp_list_of_tables = []
                temp_list_of_matches = []
                temp_qualified_teams = []
                for k in stage_group:
                    group = league_group.Group('Grupo '+str(index)+' '+str(region_sublist[counter]), k, True, self.sport, self.match_class, self.ranks)
                    index += 1
                    
                    group.generate_calendar()
                    group.simulate_league()
                    self.element_names.append(group.get_group_name())
                    table = group.get_league_table()

                    table_names = []
                    table_values = []

                    for k in table.items():
                        table_names.append(k[0])
                        table_values.append(k[1])

                    table_dict = dict(zip(table_names, table_values))
                    sorted_table = sorted(
                        table_dict.items(),
                        key= lambda item:(
                            item[1]['pts'],
                            item[1]['gd'],
                            item[1]['gf']
                        ),
                        reverse=True
                    )
                    temp_list_of_tables.append(sorted_table)
                    self.general_tables.append(sorted_table)
                    self.general_matches.append(group.get_league_matches())
                    temp_list_of_matches.append(group.get_league_matches())
                    temp_qualified_teams.append(group.get_qualified_teams(2))
                self.prev_stages_tables.append(temp_list_of_tables)
                self.prev_stages_matches.append(temp_list_of_matches)
                self.qualified_teams.append(temp_qualified_teams)
                counter += 1
                

            temp_list = []
            for qual_subgroup in self.qualified_teams:
                merged_list = list(itertools.chain.from_iterable(qual_subgroup))
                temp_list.append(merged_list)
            self.qualified_teams = temp_list
            
            index = 0
            self.group_finals = self.qualified_teams
            self.qualified_teams = []
            #Simula los torneos continentales
            for sub_qual in self.group_finals:
                tournament = tournament_group.Tournament('Copa de '+region_sublist[index]+' '+str(self.year), self.sport, sub_qual, self.ranks, False, True, self.match_class)
                trn_result = tournament.simulate_tournament()
                table = tournament.get_tournament_table()
                matches = tournament.get_tournament_matches()
                self.general_matches.append(tournament.get_tournament_matches())
                index += 1
                table_names = []
                table_values = []

                for k in table.items():
                    table_names.append(k[0])
                    table_values.append(k[1])

                table_dict = dict(zip(table_names, table_values))
                sorted_table = sorted(
                    table_dict.items(),
                    key= lambda item:(
                        item[1]['pts'],
                        item[1]['gd'],
                        item[1]['gf']
                    ),
                    reverse=True
                )
                tournament.generate_tournament_bracket(trn_result['bracket'])
                self.group_finals_tables.append(sorted_table)
                self.general_tables.append(sorted_table)
                self.group_finals_matches.append(matches)
                self.element_names.append(trn_result['tournament_name'])
                self.qualified_teams.append(tournament.get_qualified_teams(8))
            merged_list = list(itertools.chain.from_iterable(self.qualified_teams))
            self.qualified_teams = merged_list
            self.world_groups = self.subdivide_groups(8, self.qualified_teams)
            #Simula la fase de grupos mundial
            index = 1
            self.qualified_teams = []
            for w_groups in self.world_groups:
                group = league_group.Group('Grupo '+str(index)+' Mundial', w_groups, True, self.sport, self.match_class, self.ranks)
                index += 1
                group.generate_calendar()
                group.simulate_league()
                self.element_names.append(group.get_group_name())
                table = group.get_league_table()
                table_names = []
                table_values = []

                for k in table.items():
                    table_names.append(k[0])
                    table_values.append(k[1])

                table_dict = dict(zip(table_names, table_values))
                sorted_table = sorted(
                    table_dict.items(),
                    key= lambda item:(
                        item[1]['pts'],
                        item[1]['gd'],
                        item[1]['gf']
                    ),
                    reverse=True
                )
                self.general_tables.append(sorted_table)
                self.world_groups_tables.append(sorted_table)
                self.world_groups_matches.append(group.get_league_matches())
                self.general_matches.append(group.get_league_matches())
                self.qualified_teams.append(group.get_qualified_teams(2))
            merged_list = list(itertools.chain.from_iterable(self.qualified_teams))
            self.qualified_teams = merged_list
            #Simula la fase final mundial
            tournament = tournament_group.Tournament('Final '+self.sport+' '+str(self.year), self.sport, self.qualified_teams, self.ranks, False, True, self.match_class)
            trn_result = tournament.simulate_tournament()
            table = tournament.get_tournament_table()
            matches = tournament.get_tournament_matches()
            self.element_names.append(trn_result['tournament_name'])
            table_names = []
            table_values = []

            for k in table.items():
                table_names.append(k[0])
                table_values.append(k[1])

            table_dict = dict(zip(table_names, table_values))
            sorted_table = sorted(
                table_dict.items(),
                key= lambda item:(
                    item[1]['pts'],
                    item[1]['gd'],
                    item[1]['gf']
                ),
                reverse=True
            )

            tournament.generate_tournament_bracket(trn_result['bracket'])

            self.final_table = sorted_table
            self.general_tables.append(sorted_table)
            self.final_matches.append(matches)
            self.general_matches.append(matches)
            
            if self.saveres:   
                self.save_results()

            '''
            merged_list = list(itertools.chain.from_iterable(self.general_matches))
            ultra_merged = list(itertools.chain.from_iterable(merged_list))
            self.general_matches = ultra_merged

            matches_dict = [
                dict(zip(["team1","score1","team2","score2"], m))
                for m in self.general_matches
            ]
            self.general_matches = matches_dict
            '''
        else:
            if "MK" not in self.sport and "GE" in self.sport: # Simula torneos de Goldeneye #
                group_teams_sublist = []
                single_sublist = []
                for temp in self.teams:
                    if '/' in temp and 'GE' not in temp:
                        group_teams_sublist.append(temp)
                    elif 'GE' not in temp:
                        single_sublist.append(temp)
                #Individual
                if self.sport != 'GE-Teams':
                    self.prev_stages = self.subdivide_groups(self.num_groups, single_sublist)
                else:
                    self.prev_stages = self.subdivide_groups(self.num_groups, group_teams_sublist)
                print(self.prev_stages)
                #Simula los grupos de la fase previa#
                index = 1
                for stage_group in self.prev_stages:
                    group = league_group.Group('Grupo '+str(index), stage_group, True, self.sport, self.match_class, self.ranks)
                    index += 1
                    group.generate_calendar()
                    group.simulate_league()
                    self.element_names.append(group.get_group_name())
                    table = group.get_league_table()

                    table_names = []
                    table_values = []

                    for k in table.items():
                        table_names.append(k[0])
                        table_values.append(k[1])

                    table_dict = dict(zip(table_names, table_values))
                    sorted_table = sorted(
                        table_dict.items(),
                        key= lambda item:(
                            item[1]['pts'],
                            item[1]['gd'],
                            item[1]['gf']
                        ),
                        reverse=True
                    )
                    self.general_tables.append(sorted_table)
                    self.general_matches.append(group.get_league_matches())
                    self.prev_stages_tables.append(sorted_table)
                    self.prev_stages_matches.append(group.get_league_matches())
                    self.qualified_teams.append(group.get_qualified_teams(4))

                merged_list = list(itertools.chain.from_iterable(self.qualified_teams))
                self.qualified_teams = merged_list
                print(self.qualified_teams)
                #Simula la fase final mundial
                tournament = tournament_group.Tournament('Final '+self.sport+'_'+str(self.year), self.sport, self.qualified_teams, self.ranks, False, True, self.match_class)
                trn_result = tournament.simulate_tournament()
                table = tournament.get_tournament_table()
                matches = tournament.get_tournament_matches()
                self.element_names.append(trn_result['tournament_name'])
                table_names = []
                table_values = []

                for k in table.items():
                    table_names.append(k[0])
                    table_values.append(k[1])

                table_dict = dict(zip(table_names, table_values))
                sorted_table = sorted(
                    table_dict.items(),
                    key= lambda item:(
                        item[1]['pts'],
                        item[1]['gd'],
                        item[1]['gf']
                    ),
                    reverse=True
                )

                tournament.generate_tournament_bracket(trn_result['bracket'])

                self.final_table = sorted_table
                self.general_tables.append(sorted_table)
                self.final_matches.append(matches)
                self.general_matches.append(matches)
                if self.saveres:   
                    self.save_results()

            elif "MK" not in self.sport and "SSB" in self.sport: #Simula Torneos de Super Smash#
                group_teams_sublist = []
                single_sublist = []
                for temp in self.teams:
                    if '/' in temp:
                        group_teams_sublist.append(temp)
                    else:
                        single_sublist.append(temp)
                #Individual
                self.teams = single_sublist
                self.prev_stages = self.subdivide_groups(self.num_groups, self.teams)
                print(self.prev_stages)
                #Simula los grupos de la fase previa#
                index = 1
                for stage_group in self.prev_stages:
                    group = league_group.Group('Grupo '+str(index), stage_group, True, self.sport, self.match_class, self.ranks)
                    index += 1
                    group.generate_calendar()
                    group.simulate_league()
                    self.element_names.append(group.get_group_name())
                    table = group.get_league_table()

                    table_names = []
                    table_values = []

                    for k in table.items():
                        table_names.append(k[0])
                        table_values.append(k[1])

                    table_dict = dict(zip(table_names, table_values))
                    sorted_table = sorted(
                        table_dict.items(),
                        key= lambda item:(
                            item[1]['pts'],
                            item[1]['gd'],
                            item[1]['gf']
                        ),
                        reverse=True
                    )
                    self.general_tables.append(sorted_table)
                    self.general_matches.append(group.get_league_matches())
                    self.prev_stages_tables.append(sorted_table)
                    self.prev_stages_matches.append(group.get_league_matches())
                    self.qualified_teams.append(group.get_qualified_teams(1))

                merged_list = list(itertools.chain.from_iterable(self.qualified_teams))
                self.qualified_teams = merged_list
                print('Clasificados Individual')
                print(self.qualified_teams)
                #Simula la fase final mundial
                tournament = tournament_group.Tournament('Final '+self.sport+'_'+str(self.year), self.sport, self.qualified_teams, self.ranks, False, True, self.match_class)
                trn_result = tournament.simulate_tournament()
                table = tournament.get_tournament_table()
                matches = tournament.get_tournament_matches()
                self.element_names.append(trn_result['tournament_name'])
                table_names = []
                table_values = []

                for k in table.items():
                    table_names.append(k[0])
                    table_values.append(k[1])

                table_dict = dict(zip(table_names, table_values))
                sorted_table = sorted(
                    table_dict.items(),
                    key= lambda item:(
                        item[1]['pts'],
                        item[1]['gd'],
                        item[1]['gf']
                    ),
                    reverse=True
                )

                tournament.generate_tournament_bracket(trn_result['bracket'])

                self.final_table = sorted_table
                self.general_tables.append(sorted_table)
                self.final_matches.append(matches)
                self.general_matches.append(matches) 

                #Equipos
                self.prev_stages_tables = []
                self.prev_stages_matches = []
                self.qualified_teams = []
                self.final_table = {}
                self.final_matches= []
                self.teams = group_teams_sublist
                self.num_groups = 2
                self.prev_stages = self.subdivide_groups(self.num_groups, self.teams)
                print(self.prev_stages)

                #Simula los grupos de la fase previa#
                index = 1
                print(index)
                for stage_group in self.prev_stages:
                    group = league_group.Group('Grupo '+str(index), stage_group, True, self.sport, self.match_class, self.ranks)
                    index += 1
                    group.generate_calendar()
                    group.simulate_league()
                    self.element_names.append(group.get_group_name())
                    table = group.get_league_table()

                    table_names = []
                    table_values = []

                    for k in table.items():
                        table_names.append(k[0])
                        table_values.append(k[1])

                    table_dict = dict(zip(table_names, table_values))
                    sorted_table = sorted(
                        table_dict.items(),
                        key= lambda item:(
                            item[1]['pts'],
                            item[1]['gd'],
                            item[1]['gf']
                        ),
                        reverse=True
                    )
                    self.general_tables.append(sorted_table)
                    self.general_matches.append(group.get_league_matches())
                    self.prev_stages_tables.append(sorted_table)
                    self.prev_stages_matches.append(group.get_league_matches())
                    self.qualified_teams.append(group.get_qualified_teams(4))

                merged_list = list(itertools.chain.from_iterable(self.qualified_teams))
                self.qualified_teams = merged_list
                print('Clasificados equipos')
                print(self.qualified_teams)
                #Simula la fase final mundial
                tournament = tournament_group.Tournament('Final EQ '+self.sport+'_'+str(self.year), self.sport, self.qualified_teams, self.ranks, False, True, self.match_class)
                trn_result = tournament.simulate_tournament()
                table = tournament.get_tournament_table()
                matches = tournament.get_tournament_matches()
                self.element_names.append(trn_result['tournament_name'])
                table_names = []
                table_values = []

                for k in table.items():
                    table_names.append(k[0])
                    table_values.append(k[1])

                table_dict = dict(zip(table_names, table_values))
                sorted_table = sorted(
                    table_dict.items(),
                    key= lambda item:(
                        item[1]['pts'],
                        item[1]['gd'],
                        item[1]['gf']
                    ),
                    reverse=True
                )

                tournament.generate_tournament_bracket(trn_result['bracket'])

                self.final_table = sorted_table
                self.general_tables.append(sorted_table)
                self.final_matches.append(matches)
                self.general_matches.append(matches)  
                if self.saveres:   
                    self.save_results()           
            else: # Simula torneos de Mario Kart y Muñecos #
                print(self.teams)
                #Simula la fase final mundial
                tournament = tournament_group.Tournament('Final '+self.sport+'_'+str(self.year), self.sport, self.teams, self.ranks, False, True, self.match_class)
                trn_result = tournament.simulate_tournament()
                table = tournament.get_tournament_table()
                matches = tournament.get_tournament_matches()
                self.element_names.append(trn_result['tournament_name'])
                table_names = []
                table_values = []

                for k in table.items():
                    table_names.append(k[0])
                    table_values.append(k[1])

                table_dict = dict(zip(table_names, table_values))
                sorted_table = sorted(
                    table_dict.items(),
                    key= lambda item:(
                        item[1]['pts'],
                        item[1]['gd'],
                        item[1]['gf']
                    ),
                    reverse=True
                )

                tournament.generate_tournament_bracket(trn_result['bracket'])

                self.final_table = sorted_table
                self.general_tables.append(sorted_table)
                self.final_matches.append(matches)
                self.general_matches.append(matches)
                if self.saveres:   
                    self.save_results()

        pass


    def save_results(self):
        if self.match_class in [1,5]:
            merged_table = self.merge_tables(self.general_tables)
            print(merged_table)
            print(len(merged_table))
            index = 1
            for eq in merged_table:
                team_obj = Nationalteams.objects.get(team_name = eq[0])
                sport = Teamsports.objects.get(team_sport_name = self.sport)
                try:
                    existing_log = Teamtournamentregister.objects.get(team_id = team_obj.team_id, team_year = str(self.year), team_sport_id = sport.team_sport_id)
                    existing_log.team_id = team_obj.team_id
                    existing_log.team_wins = eq[1]['w']
                    existing_log.team_draws = eq[1]['d']
                    existing_log.team_loses = eq[1]['l']
                    existing_log.team_sc_points = eq[1]['gf']
                    existing_log.team_ag_points = eq[1]['gc']
                    existing_log.team_position = index
                    existing_log.team_year = str(self.year)
                    existing_log.team_sport_id = sport.team_sport_id
                    existing_log.save()

                except Teamtournamentregister.DoesNotExist:
                    tournament_element = Teamtournamentregister(
                        team_id = team_obj.team_id,
                        team_wins = eq[1]['w'],
                        team_draws = eq[1]['d'],
                        team_loses = eq[1]['l'],
                        team_sc_points = eq[1]['gf'],
                        team_ag_points = eq[1]['gc'],
                        team_position = index,
                        team_year = str(self.year),
                        team_sport_id = sport.team_sport_id
                    )
                    tournament_element.save()
                index += 1

            sport = Teamsports.objects.get(team_sport_name = self.sport)
            Teammatchesregister.objects.filter(match_year = str(self.year), team_sport_id = sport.team_sport_id).delete()

            for cont in self.general_matches:
                for m in cont:
                    #print(m['team1'], m['team2'])
                    team1_obj = Nationalteams.objects.get(team_name = m['team1'])
                    team2_obj = Nationalteams.objects.get(team_name = m['team2'])
                    sport = Teamsports.objects.get(team_sport_name = self.sport)
                    result_label = ''

                    if int(m['score1']) > int(m['score2']):
                        result_label = m['team1'] + ' W.'
                    elif int(m['score2']) > int(m['score1']):
                        result_label = m['team2'] + ' W.'
                    else:
                        result_label = 'D.'
                

                    match_element = Teammatchesregister(
                        team_local_id = team1_obj.team_id,
                        team_local_score = m['score1'],
                        team_away_id = team2_obj.team_id,
                        team_away_score = m['score2'],
                        result_label = result_label,
                        team_sport_id = sport.team_sport_id,
                        match_year = str(self.year)
                    )
                    match_element.save()
                

            pass
        elif self.match_class in [2,4]:
            merged_table = self.merge_tables(self.general_tables)
            print(merged_table)
            print(len(merged_table))
            index = 1
            for eq in merged_table:
                player_obj = Olympicplayers.objects.get(ol_player_name = eq[0])
                sport = Playertournamentsports.objects.get(player_trn_sport_name = self.sport)
                try:
                    existing_log = Playertournamentregister.objects.get(ol_player_id = player_obj.ol_player_id, ol_player_year = str(self.year), player_trn_sport_id = sport.player_trn_sport_id)
                    existing_log.ol_player_id = player_obj.ol_player_id
                    existing_log.ol_player_wins = eq[1]['w']
                    existing_log.ol_player_draws = eq[1]['d']
                    existing_log.ol_player_loses = eq[1]['l']
                    existing_log.ol_player_sc_points = eq[1]['gf']
                    existing_log.ol_player_ag_points = eq[1]['gc']
                    existing_log.ol_player_position = index
                    existing_log.ol_player_year = str(self.year)
                    existing_log.player_trn_sport_id = sport.player_trn_sport_id
                    existing_log.save()

                except Playertournamentregister.DoesNotExist:
                    tournament_element = Playertournamentregister(
                        ol_player_id = player_obj.ol_player_id,
                        ol_player_wins = eq[1]['w'],
                        ol_player_draws = eq[1]['d'],
                        ol_player_loses = eq[1]['l'],
                        ol_player_sc_points = eq[1]['gf'],
                        ol_player_ag_points = eq[1]['gc'],
                        ol_player_position = index,
                        ol_player_year = str(self.year),
                        player_trn_sport_id = sport.player_trn_sport_id
                    )
                    tournament_element.save()
                index += 1
        pass



    def merge_tables(self, tables):
        merged = defaultdict(lambda: {
            "pts": 0,
            "w": 0,
            "l": 0,
            "d": 0,
            "gf": 0,
            "gc": 0,
            "gd": 0
        })

        for table in tables:              # cada "tabla"
            for team, stats in table:     # cada tupla ("team", {...})
                for key, value in stats.items():
                    merged[team][key] += value

        # Convertir al formato original: lista de tuplas
        result = [(team, stats) for team, stats in merged.items()]
        result.sort(key=lambda x: (x[1]["pts"], x[1]["gd"], x[1]["gf"]), reverse=True)
        return result

    def generate_tournament_excel(self, file_path="tournament_simulation.xlsx"):
        merged_tables = self.merge_tables(self.general_tables)
        self.general_tables.append(merged_tables)
        self.element_names.append('Tabla General')
        self.general_matches.append([])
        wb = Workbook()
        wb.remove(wb.active)

        for i in range(len(self.element_names)):

            sheet_name = self.element_names[i][:31]  # Excel limite nombre hoja
            ws = wb.create_sheet(title=sheet_name)

            table = self.general_tables[i]
            matches = self.general_matches[i]

            # -------- TABLA --------
            ws["A1"] = "Tabla"
            ws["A1"].font = Font(bold=True)

            headers = ["Pos", "Team", "Pts", "W", "D", "L", "GF", "GC", "GD"]

            for col, header in enumerate(headers, start=1):
                ws.cell(row=2, column=col, value=header).font = Font(bold=True)

            row = 3
            pos = 1

            for team, stats in table:

                ws.cell(row=row, column=1, value=pos)
                ws.cell(row=row, column=2, value=team)
                ws.cell(row=row, column=3, value=stats["pts"])
                ws.cell(row=row, column=4, value=stats["w"])
                ws.cell(row=row, column=5, value=stats["d"])
                ws.cell(row=row, column=6, value=stats["l"])
                ws.cell(row=row, column=7, value=stats["gf"])
                ws.cell(row=row, column=8, value=stats["gc"])
                ws.cell(row=row, column=9, value=stats["gd"])

                row += 1
                pos += 1

            # -------- PARTIDOS --------
            start_row = row + 2

            ws.cell(row=start_row, column=1, value="Partidos").font = Font(bold=True)

            match_headers = ["Team 1", "Team 2", "Score 1", "Score 2"]

            for col, header in enumerate(match_headers, start=1):
                ws.cell(row=start_row + 1, column=col, value=header).font = Font(bold=True)

            r = start_row + 2
            if isinstance(matches, dict):
                matches = [matches]
            for match in matches:

                for col, key in enumerate(["team1","team2","score1","score2"], start=1):
                    ws.cell(row=r, column=col, value=match[key])

                r += 1

        wb.save(file_path)

        return file_path

    
