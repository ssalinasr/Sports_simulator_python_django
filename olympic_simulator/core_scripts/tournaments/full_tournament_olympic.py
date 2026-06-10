from core_scripts.leagues import league_group
from core_scripts.tournaments import tournament_group
from core_scripts.leagues import league_tools
import random, re
import itertools
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from appolympics.models import Nationalteams, Olympicplayers, Teamsports, Teammedalregister, Teamsimulationregister, Playersimulationregister, Playermedalregister
from appolympics.models import Playertitleregister, Teamtitleregister, Sportsrecords
from core_scripts.interfaces.olympic_sports_interfaces import simulated_sports

class FullTournament():

    def __init__(self, simulation_pairs, competence, num_groups, match_class, year, saveres):
        self.simulation_pairs = simulation_pairs
        self.match_class = match_class
        self.competence = competence
        self.year = year
        self.saveres = saveres
        self.element_names = []
        self.num_groups = num_groups
        self.general_tables = []
        self.medalists = []
        self.save_gen_tables = []

    def subdivide_groups(self, num_groups, teams):
        random.shuffle(teams)
        base_length = len(teams) // num_groups
        remainder = len(teams) % num_groups

        sorted_teams = []
        index = 0

        for i in range(num_groups):
            length = base_length + 1 if i < remainder else base_length
            sorted_teams.append(teams[index:index+length])
            index += length
        
        return sorted_teams

    def simulate_tournament(self):
        if self.match_class == 1:
            qual_list = [64,32,16,8]
            for q_num in qual_list:
                for sp in self.simulation_pairs:
                    #print('Categoria: ',sp[0], ' Num participantes: ',len(sp[2]), ' Deportes: ',sp[1])
                    for dep in sp[1]:
                        sim_sport = simulated_sports.SimulatedSports(sp[0],dep,[],sp[2], self.match_class)
                        sim_sport.simulate_olympic_sport()
                        #print('Este es el deporte ', dep.sp_record_name)
                        table_res = sim_sport.get_table_results()
                        new_table = defaultdict(lambda: {
                            "ranking": 0,
                            "score": 0
                        })
                        num_qual = 0
                        qual_teams = []
                        for tr in table_res:
                            if num_qual < q_num:
                                qual_teams.append((tr[0],tr[1]))
                                num_qual += 1
                            new_table[tr[0]]["ranking"] = tr[1]
                            try:
                                new_table[tr[0]]["score"] = tr[2][1]
                            except:
                                new_table[tr[0]]["score"] = tr[2]

                        result_table = [(team, stats) for team, stats in new_table.items()]

                        self.general_tables.append(result_table)
                        self.save_gen_tables.append((dep.sp_record_name, sp[0], result_table))
                        if q_num == 64:
                            self.element_names.append(dep.sp_record_name + ' Q')
                        elif q_num == 32:
                            self.element_names.append(dep.sp_record_name + ' CF')
                        elif q_num == 16:
                            self.element_names.append(dep.sp_record_name + ' SF')
                        elif q_num == 8:
                            self.element_names.append(dep.sp_record_name + ' F')


                        #print('Clasificados:', len(qual_teams))
                        sp[2] = qual_teams

                        if q_num == 8:
                            medal_winners = list(new_table.items())[:3]
                            medalist_list = []
                            for m in medal_winners:
                                medalist_list.append(m[0])
                            self.medalists.append((dep.sp_record_name, sp[0], medalist_list))
                   
            pass
        elif self.match_class in [2,4]:
            for sp in self.simulation_pairs:
                #print('Categoria: ',sp[0], ' Num participantes: ',len(sp[2]), ' Deportes: ',sp[1])
                for dep in sp[1]:
                    sim_sport = simulated_sports.SimulatedSports(sp[0],dep,[],sp[2], self.match_class)
                    sim_sport.simulate_olympic_sport()
                    #print('Este es el deporte ', dep.sp_record_name)
                    table_res = sim_sport.get_table_results()
                    new_table = defaultdict(lambda: {
                        "ranking": 0,
                        "score": 0
                    })
                    for tr in table_res:
                        new_table[tr[0]]["ranking"] = tr[1]
                        try:
                            new_table[tr[0]]["score"] = tr[2][1]
                        except:
                            new_table[tr[0]]["score"] = tr[2]

                    result_table = [(team, stats) for team, stats in new_table.items()]
                    self.general_tables.append(result_table)
                    self.save_gen_tables.append((dep.sp_record_name, sp[0], result_table))
                    self.element_names.append(dep.sp_record_name)
                    #print(new_table)
                    medal_winners = list(new_table.items())[:3]
                    medalist_list = []
                    for m in medal_winners:
                        medalist_list.append(m[0])
                    self.medalists.append((dep.sp_record_name,sp[0], medalist_list))
                         
            pass

        if self.saveres:   
            self.save_results()
        pass


    def save_results(self):
        if self.match_class in [1,5]:
            for mt in self.save_gen_tables:
                for eq in mt[2]:
                    team_obj = Nationalteams.objects.get(team_name = eq[0])
                    sport = Sportsrecords.objects.get(sp_record_name = mt[0], team_sport_id = mt[1])
                    try:
                        existing_log = Teamsimulationregister.objects.get(team_id = team_obj.team_id, team_year = str(self.year), sp_record = sport.sp_record_id)
                        existing_log.team_id = team_obj.team_id
                        if isinstance(eq[1]["score"], list):
                            existing_log.team_result = eq[1]["score"][0]
                        else:
                            existing_log.team_result = eq[1]["score"]
                        existing_log.team_year = str(self.year)
                        existing_log.sp_record = sport
                        existing_log.save()

                    except Teamsimulationregister.DoesNotExist:
                        if isinstance(eq[1]["score"], list):
                            team_result = eq[1]["score"][0]
                        else:
                            team_result = eq[1]["score"]

                        tournament_element = Teamsimulationregister(
                            team_id = team_obj.team_id,
                            team_result = team_result,
                            team_year = str(self.year),
                            sp_record = sport
                        )
                        tournament_element.save()

            for medal in self.medalists:    
                sport = Sportsrecords.objects.get(sp_record_name = medal[0], team_sport_id = medal[1])
                counter = 0
                for md in medal[2]:
                    team_obj = Nationalteams.objects.get(team_name = md)
                    try:
                        existing_log = Teammedalregister.objects.get(team_id = team_obj.team_id, sp_record_id = sport.sp_record_id, medal_year = str(self.year))
                        existing_log.team_id = team_obj.team_id
                        if counter == 0:
                            existing_log.medal_label = 'O'
                        elif counter == 1:
                            existing_log.medal_label = 'P'
                        elif counter == 2:
                            existing_log.medal_label = 'B'
                        existing_log.medal_year = str(self.year)
                        existing_log.sp_record = sport
                        existing_log.save()

                    except Teammedalregister.DoesNotExist:
                        if counter == 0:
                            title_label = 'O'
                        elif counter == 1:
                            title_label = 'P'
                        elif counter == 2:
                            title_label = 'B'
                        title_element = Teammedalregister(
                            team_id = team_obj.team_id,
                            medal_label = title_label,
                            medal_year = str(self.year),
                            sp_record_id = sport.sp_record_id
                        )
                        title_element.save()
                        counter += 1
                pass
            
    
        elif self.match_class in [2,4]:
            for mt in self.save_gen_tables:
                for eq in mt[2]:
                    print(eq[0])
                    team_obj = Olympicplayers.objects.filter(ol_player_name = eq[0]).first()
                    sport = Sportsrecords.objects.get(sp_record_name = mt[0], team_sport_id = mt[1])
                    try:
                        existing_log = Playersimulationregister.objects.get(ol_player_id = team_obj.ol_player_id, ol_player_year = str(self.year), sp_record = sport.sp_record_id)
                        existing_log.ol_player_id = team_obj.ol_player_id
                        if isinstance(eq[1]["score"], list):
                            existing_log.ol_player_result = eq[1]["score"][0]
                        else:
                            existing_log.ol_player_result = eq[1]["score"]
                        existing_log.ol_player_year = str(self.year)
                        existing_log.sp_record = sport
                        existing_log.save()

                    except Playersimulationregister.DoesNotExist:
                        if isinstance(eq[1]["score"], list):
                            team_result = eq[1]["score"][0]
                        else:
                            team_result = eq[1]["score"]
                        tournament_element = Playersimulationregister(
                            ol_player_id = team_obj.ol_player_id,
                            ol_player_result = team_result,
                            ol_player_year = str(self.year),
                            sp_record = sport
                        )
                        tournament_element.save()
            for medal in self.medalists:    
                sport = Sportsrecords.objects.get(sp_record_name = medal[0], team_sport_id = medal[1])
                counter = 0
                for md in medal[2]:
                    team_obj = Olympicplayers.objects.filter(ol_player_name = md).first()
                    try:
                        existing_log = Playermedalregister.objects.get(ol_player_id = team_obj.ol_player_id, sp_record_id = sport.sp_record_id, medal_year = str(self.year))
                        existing_log.ol_player_id = team_obj.ol_player_id
                        if counter == 0:
                            existing_log.medal_label = 'O'
                        elif counter == 1:
                            existing_log.medal_label = 'P'
                        elif counter == 2:
                            existing_log.medal_label = 'B'
                        existing_log.medal_year = str(self.year)
                        existing_log.sp_record = sport
                        existing_log.save()

                    except Playermedalregister.DoesNotExist:
                        if counter == 0:
                            title_label = 'O'
                        elif counter == 1:
                            title_label = 'P'
                        elif counter == 2:
                            title_label = 'B'
                        title_element = Playermedalregister(
                            ol_player_id = team_obj.ol_player_id,
                            medal_label = title_label,
                            medal_year = str(self.year),
                            sp_record_id = sport.sp_record_id
                        )
                        title_element.save()
                        counter += 1
    pass

    def merge_tables(self, tables):
        merged = defaultdict(lambda: {
            "pts": 0,
            "w": 0,
            "l": 0,
            "d": 0,
            "gf": 0,
            "gc": 0,
            "gd": 0
        })

        for table in tables:              # cada "tabla"
            for team, stats in table:     # cada tupla ("team", {...})
                for key, value in stats.items():
                    merged[team][key] += value

        # Convertir al formato original: lista de tuplas
        result = [(team, stats) for team, stats in merged.items()]
        result.sort(key=lambda x: (x[1]["pts"], x[1]["gd"], x[1]["gf"]), reverse=True)
        return result

    def generate_tournament_excel(self, file_path="tournament_simulation.xlsx"):
        wb = Workbook()
        wb.remove(wb.active)

        for i in range(len(self.element_names)):

            sheet_name = f"Hoja_{i+1:03d}"

            ws = wb.create_sheet(title=sheet_name)

            table = self.general_tables[i]
            # -------- TABLA --------
            ws.merge_cells("A1:D1")
            ws["A1"] = self.element_names[i]
            ws["A1"].font = Font(bold=True)
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

            headers = ["Pos", "Team", "Ranking", "Score"]

            for col, header in enumerate(headers, start=1):
                ws.cell(row=2, column=col, value=header).font = Font(bold=True)

            row = 3
            pos = 1


            for team, stats in table:
                ws.cell(row=row, column=1, value=pos)
                ws.cell(row=row, column=2, value=team)
                ws.cell(row=row, column=3, value=stats["ranking"])
                try:
                    ws.cell(row=row, column=4, value=stats["score"])
                except:
                    ws.cell(row=row, column=4, value=stats["score"][0])

                row += 1
                pos += 1

        wb.save(file_path)

        return file_path

    
