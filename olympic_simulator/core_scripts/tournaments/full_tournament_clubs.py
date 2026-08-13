from core_scripts.leagues import league_group
from core_scripts.tournaments import tournament_group
from core_scripts.leagues import league_tools
import random
import itertools
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse
from appolympics.models import Clubs, Clubmatchesregister, Clubtournamentregister, Clubtitleregister, Clubmedalregister, Sportsrecords
from appolympics.models import Teammedalregister, Nationalteams
class FullTournamentClubs():

    def __init__(self, teams, league_name, ranks, sport, num_groups, match_class, year, saveres):
        self.teams = teams
        self.league_name = league_name
        self.ranks = ranks
        self.sport = sport
        self.match_class = match_class
        self.year = year
        self.saveres = saveres
        self.world_groups = []
        self.world_groups_tables = []
        self.world_groups_matches = []
        self.qualified_teams = []
        self.final = []
        self.final_table = {}
        self.final_matches = []
        self.element_names = []
        self.num_groups = num_groups
        self.general_tables = []
        self.general_matches = []
        self.world_qualified = []
        self.champions = []

    def subdivide_groups(self, num_groups, teams):
        random.shuffle(teams)
        base_length = len(teams) // num_groups
        remainder = len(teams) % num_groups

        sorted_teams = []
        index = 0

        for i in range(num_groups):
            length = base_length + 1 if i < remainder else base_length
            sorted_teams.append(teams[index:index+length])
            index += length
        
        return sorted_teams

    def simulate_tournament(self):
        #Simula la fase de grupos mundial
        index = 1
        self.qualified_teams = []

        if self.num_groups > 1:
            self.world_groups = self.subdivide_groups(self.num_groups, self.teams)
        else:
            self.world_groups = self.subdivide_groups(self.num_groups, self.teams)
        
        for w_groups in self.world_groups:
            if self.league_name in 'UEFA':
                group = league_group.Group(self.league_name + ' '+ str(index), w_groups, False, self.sport, self.match_class, self.ranks)
            else:
                group = league_group.Group(self.league_name + ' '+ str(index), w_groups, True, self.sport, self.match_class, self.ranks)
            index += 1
            group.generate_calendar()
            group.simulate_league()
            self.element_names.append(group.get_group_name())
            table = group.get_league_table()
            table_names = []
            table_values = []

            for k in table.items():
                table_names.append(k[0])
                table_values.append(k[1])

            table_dict = dict(zip(table_names, table_values))
            sorted_table = sorted(
                table_dict.items(),
                key= lambda item:(
                    item[1]['pts'],
                    item[1]['gd'],
                    item[1]['gf']
                ),
                reverse=True
            )
            self.general_tables.append(sorted_table)
            self.world_groups_tables.append(sorted_table)
            self.world_groups_matches.append(group.get_league_matches())
            self.general_matches.append(group.get_league_matches())
            if self.num_groups > 1:
                self.qualified_teams.append(group.get_qualified_teams(2))
            else:
                self.qualified_teams.append(group.get_qualified_teams(16))
                
        merged_list = list(itertools.chain.from_iterable(self.qualified_teams))
        self.qualified_teams = merged_list
        #Simula la fase final mundial
        tournament = tournament_group.Tournament('Final '+self.league_name+' '+str(self.year), self.sport, self.qualified_teams, self.ranks, False, True, self.match_class)
        trn_result = tournament.simulate_tournament()
        table = tournament.get_tournament_table()
        matches = tournament.get_tournament_matches()
        self.element_names.append(trn_result['tournament_name'])

        img = tournament.generate_tournament_bracket(trn_result['bracket'])
        self.champions.append((trn_result['champion'], trn_result['tournament_name'], trn_result['bracket'], img))
        table_names = []
        table_values = []

        for k in table.items():
            table_names.append(k[0])
            table_values.append(k[1])

        table_dict = dict(zip(table_names, table_values))
        sorted_table = sorted(
            table_dict.items(),
            key= lambda item:(
                item[1]['pts'],
                item[1]['gd'],
                item[1]['gf']
            ),
            reverse=True
        )

        
        if self.league_name in ['UEFA Champions League','Copa Libertadores', 'CAF Champions League', 'AFC Champions League Elite']:
           self.world_qualified.append(tournament.get_qualified_teams(8))

        self.final_table = sorted_table
        self.general_tables.append(sorted_table)
        self.final_matches.append(matches)
        self.general_matches.append(matches)
        
        if self.saveres:   
            self.save_results()
        pass

    def get_world_qualified(self):
        return self.world_qualified

    def save_results(self):
        merged_table = self.merge_tables(self.general_tables)
        index = 1
        for eq in merged_table:
            club_obj = Clubs.objects.get(club_name = eq[0])
            try:
                existing_log = Clubtournamentregister.objects.get(club_id = club_obj.club_id, club_year = str(self.year), club_trn = self.league_name)
                existing_log.club_id = club_obj.club_id
                existing_log.club_wins = eq[1]['w']
                existing_log.club_draws = eq[1]['d']
                existing_log.club_loses = eq[1]['l']
                existing_log.club_sc_points = eq[1]['gf']
                existing_log.club_ag_points = eq[1]['gc']
                existing_log.club_position = index
                existing_log.club_year = str(self.year)
                existing_log.club_trn = self.league_name
                existing_log.save()

            except Clubtournamentregister.DoesNotExist:
                tournament_element = Clubtournamentregister(
                    club_id = club_obj.club_id,
                    club_wins = eq[1]['w'],
                    club_draws = eq[1]['d'],
                    club_loses = eq[1]['l'],
                    club_sc_points = eq[1]['gf'],
                    club_ag_points = eq[1]['gc'],
                    club_position = index,
                    club_year = str(self.year),
                    club_trn = self.league_name
                )
                tournament_element.save()
            index += 1

        for title in self.champions:
            try:
                club_obj = Clubs.objects.get(club_name = title[0])
                try:
                    existing_log = Clubtitleregister.objects.get(club_id = club_obj.club_id, title_year = str(self.year), title_label=title[1])
                    existing_log.club_id = club_obj.club_id
                    existing_log.title_label = title[1]
                    existing_log.title_year = str(self.year)
                    existing_log.title_bracket = dict(title[2])
                    existing_log.title_image = title[3]
                    existing_log.save()

                except Clubtitleregister.DoesNotExist:
                    title_element = Clubtitleregister(
                        club_id = club_obj.club_id,
                        title_label = title[1],
                        title_year = str(self.year),
                        title_bracket = dict(title[2]),
                        title_image = title[3]
                    )
                    title_element.save()

                    print(
                        "¿Existe después del save?:",
                        Clubtitleregister.objects.filter(
                            club_title_id = title_element.club_title_id
                        ).exists()
                    )

                champion = ''
                not_champion = ''
                element = title[2]['Final'][0]
                if element['winner'] == element['team1']:
                    champion = element['team1']
                    not_champion = element['team2']
                else:
                    champion = element['team2']
                    not_champion = element['team1']
                third_place = title[2]['Third Place'][0]['winner']

                name_sp = "Clubes - " + title[1].replace(self.year, '').replace(self.sport,'').replace('Final ','').rstrip()
                team_obj = Clubs.objects.get(club_name = champion)
                sport = Sportsrecords.objects.get(sp_record_name = name_sp)                
                try:
                    existing_log = Clubmedalregister.objects.get(club_id = team_obj.club_id, sp_record_id = sport.sp_record_id, medal_year = self.year)
                    existing_log.club_id = team_obj.club_id
                    existing_log.medal_label = 'O'
                    existing_log.medal_year = self.year
                    existing_log.sp_record = sport
                    existing_log.save()
                except Clubmedalregister.DoesNotExist:
                    title_label = 'O'
                    title_element = Clubmedalregister(
                        club_id = team_obj.club_id,
                        medal_label = title_label,
                        medal_year = self.year,
                        sp_record_id = sport.sp_record_id
                    )
                    title_element.save()
                team_obj = Clubs.objects.get(club_name = not_champion)
                sport = Sportsrecords.objects.get(sp_record_name = name_sp)      
                try:
                    existing_log = Clubmedalregister.objects.get(club_id = team_obj.club_id, sp_record_id = sport.sp_record_id, medal_year = self.year)
                    existing_log.club_id = team_obj.club_id
                    existing_log.medal_label = 'P'
                    existing_log.medal_year = self.year
                    existing_log.sp_record = sport
                    existing_log.save()
                except Clubmedalregister.DoesNotExist:
                    title_label = 'P'
                    title_element = Clubmedalregister(
                        club_id = team_obj.club_id,
                        medal_label = title_label,
                        medal_year = self.year,
                        sp_record_id = sport.sp_record_id
                    )
                    title_element.save()
                team_obj = Clubs.objects.get(club_name = third_place)
                sport = Sportsrecords.objects.get(sp_record_name = name_sp)      
                try:
                    existing_log = Clubmedalregister.objects.get(club_id = team_obj.club_id, sp_record_id = sport.sp_record_id, medal_year = self.year)
                    existing_log.club_id = team_obj.club_id
                    existing_log.medal_label = 'B'
                    existing_log.medal_year = self.year
                    existing_log.sp_record = sport
                    existing_log.save()
                except Clubmedalregister.DoesNotExist:
                    title_label = 'B'
                    title_element = Clubmedalregister(
                        club_id = team_obj.club_id,
                        medal_label = title_label,
                        medal_year = self.year,
                        sp_record_id = sport.sp_record_id
                    )
                    title_element.save()
                club_obj = Clubs.objects.get(club_name = champion)
                team_obj = Nationalteams.objects.get(team_name = club_obj.club_country.team_name)
                sport = Sportsrecords.objects.get(sp_record_name = name_sp)                
                try:
                    existing_log = Teammedalregister.objects.get(team_id = team_obj.team_id, sp_record_id = sport.sp_record_id, medal_year = self.year)
                    existing_log.team_id = team_obj.team_id
                    existing_log.medal_label = 'O'
                    existing_log.medal_year = self.year
                    existing_log.sp_record = sport
                    existing_log.save()
                except Teammedalregister.DoesNotExist:
                    title_label = 'O'
                    title_element = Teammedalregister(
                        team_id = team_obj.team_id,
                        medal_label = title_label,
                        medal_year = self.year,
                        sp_record_id = sport.sp_record_id
                    )
                    title_element.save()
                club_obj = Clubs.objects.get(club_name = not_champion)
                team_obj = Nationalteams.objects.get(team_name = club_obj.club_country.team_name)
                sport = Sportsrecords.objects.get(sp_record_name = name_sp)      
                try:
                    existing_log = Teammedalregister.objects.get(team_id = team_obj.team_id, sp_record_id = sport.sp_record_id, medal_year = self.year)
                    existing_log.team_id = team_obj.team_id
                    existing_log.medal_label = 'P'
                    existing_log.medal_year = self.year
                    existing_log.sp_record = sport
                    existing_log.save()
                except Teammedalregister.DoesNotExist:
                    title_label = 'P'
                    title_element = Teammedalregister(
                        team_id = team_obj.team_id,
                        medal_label = title_label,
                        medal_year = self.year,
                        sp_record_id = sport.sp_record_id
                    )
                    title_element.save()
                club_obj = Clubs.objects.get(club_name = third_place)
                team_obj = Nationalteams.objects.get(team_name = club_obj.club_country.team_name)
                sport = Sportsrecords.objects.get(sp_record_name = name_sp)      
                try:
                    existing_log = Teammedalregister.objects.get(team_id = team_obj.team_id, sp_record_id = sport.sp_record_id, medal_year = self.year)
                    existing_log.team_id = team_obj.team_id
                    existing_log.medal_label = 'B'
                    existing_log.medal_year = self.year
                    existing_log.sp_record = sport
                    existing_log.save()
                except Teammedalregister.DoesNotExist:
                    title_label = 'B'
                    title_element = Teammedalregister(
                        team_id = team_obj.team_id,
                        medal_label = title_label,
                        medal_year = self.year,
                        sp_record_id = sport.sp_record_id
                    )
                    title_element.save()
            except Exception as e:
                import traceback

                print(f"\n========== ERROR: {title[0]} ==========")
                print(f"Tipo: {type(e).__name__}")
                print(f"Mensaje: {e}")
                traceback.print_exc()

                raise


        registros = Clubtitleregister.objects.filter(
            title_year=str(self.year)
        )
        print("\n========== VERIFICACIÓN FINAL ==========")

        for registro in registros:
            print(
                f"ID={registro.club_title_id} | "
                f"Club={registro.club_id} | "
                f"Año={registro.title_year} | "
                f"Label={registro.title_label}"
            )

        print("TOTAL:", registros.count())
        print("CHAMPIONS GENERADOS:", len(self.champions))

        for cont in self.general_matches:
            for m in cont:
                club1_obj = Clubs.objects.get(club_name = m['team1'])
                club2_obj = Clubs.objects.get(club_name = m['team2'])
                result_label = ''

                if int(m['score1']) > int(m['score2']):
                    result_label = m['team1'] + ' W.'
                elif int(m['score2']) > int(m['score1']):
                    result_label = m['team2'] + ' W.'
                else:
                    result_label = 'D.'
            

                match_element = Clubmatchesregister(
                    club_local_id = club1_obj.club_id,
                    club_local_score = m['score1'],
                    club_away_id = club2_obj.club_id,
                    club_away_score = m['score2'],
                    result_label = result_label,
                    match_year = str(self.year)
                )
                match_element.save()




    def merge_tables(self, tables):
        merged = defaultdict(lambda: {
            "pts": 0,
            "w": 0,
            "l": 0,
            "d": 0,
            "gf": 0,
            "gc": 0,
            "gd": 0
        })

        for table in tables:              # cada "tabla"
            for team, stats in table:     # cada tupla ("team", {...})
                for key, value in stats.items():
                    merged[team][key] += value

        # Convertir al formato original: lista de tuplas
        result = [(team, stats) for team, stats in merged.items()]
        result.sort(key=lambda x: (x[1]["pts"], x[1]["gd"], x[1]["gf"]), reverse=True)
        return result

    def generate_tournament_excel(self, file_path="tournament_simulation.xlsx"):
        merged_tables = self.merge_tables(self.general_tables)
        self.general_tables.append(merged_tables)
        self.element_names.append('Tabla General')
        self.general_matches.append([])
        wb = Workbook()
        wb.remove(wb.active)

        for i in range(len(self.element_names)):

            sheet_name = self.element_names[i][:31]  # Excel limite nombre hoja
            ws = wb.create_sheet(title=sheet_name)

            table = self.general_tables[i]
            matches = self.general_matches[i]

            # -------- TABLA --------
            ws["A1"] = "Tabla"
            ws["A1"].font = Font(bold=True)

            headers = ["Pos", "Team", "Pts", "W", "D", "L", "GF", "GC", "GD"]

            for col, header in enumerate(headers, start=1):
                ws.cell(row=2, column=col, value=header).font = Font(bold=True)

            row = 3
            pos = 1

            for team, stats in table:

                ws.cell(row=row, column=1, value=pos)
                ws.cell(row=row, column=2, value=team)
                ws.cell(row=row, column=3, value=stats["pts"])
                ws.cell(row=row, column=4, value=stats["w"])
                ws.cell(row=row, column=5, value=stats["d"])
                ws.cell(row=row, column=6, value=stats["l"])
                ws.cell(row=row, column=7, value=stats["gf"])
                ws.cell(row=row, column=8, value=stats["gc"])
                ws.cell(row=row, column=9, value=stats["gd"])

                row += 1
                pos += 1

            # -------- PARTIDOS --------
            start_row = row + 2

            ws.cell(row=start_row, column=1, value="Partidos").font = Font(bold=True)

            match_headers = ["Team 1", "Team 2", "Score 1", "Score 2"]

            for col, header in enumerate(match_headers, start=1):
                ws.cell(row=start_row + 1, column=col, value=header).font = Font(bold=True)

            r = start_row + 2
            if isinstance(matches, dict):
                matches = [matches]
            for match in matches:

                for col, key in enumerate(["team1","team2","score1","score2"], start=1):
                    ws.cell(row=r, column=col, value=match[key])

                r += 1

        wb.save(file_path)

        return file_path

    
