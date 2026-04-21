from appolympics.models import Clubs, Clubleague
from core_scripts.leagues import league_group
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font

class ClubLeagueSeason:
    def __init__(self, teams_tuple, country, has_promotion, year, promotions, qualifiers, region, has_save, ranks):
        self.teams_tuple = teams_tuple
        self.country = country
        self.has_promotion = has_promotion
        self.year = year
        self.has_save = has_save
        self.ranks = ranks
        self.promotions = promotions
        self.qualifiers = qualifiers
        self.region = region
        self.first_cup_qualified = []
        self.first_cup_prev_qualified = []
        self.second_cup_qualified = []
        self.second_cup_prev_qualified = []
        self.third_cup_qualified = []
        self.third_cup_prev_qualified = []
        self.promoted_teams = []
        self.relegated_teams = []
        self.general_matches = []
        self.general_tables = []
        self.element_names = []


    def simulate_league(self):
        groups = league_group.Group(self.teams_tuple[2], self.teams_tuple[0], True, 'Futbol Masculino', 3, self.ranks)  
        groups.generate_calendar()
        groups.simulate_league()
        self.element_names.append(groups.get_group_name())
        table = groups.get_league_table()
        matches = groups.get_league_matches()
        
        table_names = []
        table_values = []

        for k in table.items():
            table_names.append(k[0])
            table_values.append(k[1])

        table_dict = dict(zip(table_names, table_values))
        sorted_table = sorted(
            table_dict.items(),
            key= lambda item:(
                item[1]['pts'],
                item[1]['gd'],
                item[1]['gf']
            ),
            reverse=True
        )

        self.general_matches.append(matches)
        self.general_tables.append(sorted_table)

        if self.teams_tuple[3] == '1D':
            if self.has_promotion:
                self.relegated_teams.append(groups.get_qualified_reversed_teams(self.promotions))
            else:
                pass

            if self.teams_tuple[7] == 'Y':
                self.first_cup_qualified.append(groups.get_qualified_specified_teams(0, self.qualifiers[0]))
                self.second_cup_qualified.append(groups.get_qualified_specified_teams(self.qualifiers[0], self.qualifiers[0]+self.qualifiers[1]))
                self.third_cup_qualified.append(groups.get_qualified_specified_teams(self.qualifiers[0]+self.qualifiers[1], self.qualifiers[0]+self.qualifiers[1]+self.qualifiers[2]))
            else:
                self.first_cup_prev_qualified.append(groups.get_qualified_specified_teams(0, self.qualifiers[0]))
                self.second_cup_prev_qualified.append(groups.get_qualified_specified_teams(self.qualifiers[0], self.qualifiers[0]+self.qualifiers[1]))
                self.third_cup_prev_qualified.append(groups.get_qualified_specified_teams(self.qualifiers[0]+self.qualifiers[1], self.qualifiers[0]+self.qualifiers[1]+self.qualifiers[2]))
        elif self.teams_tuple[3] == '2D':
            self.promoted_teams.append(groups.get_qualified_specified_teams(0, self.promotions))

    def update_promotions_relegations(self):
        pass

    def get_full_results(self):
        first_cup = self.get_first_cup_qualified_teams()
        first_cup_prev = self.get_first_cup_prev_qualified_teams()
        second_cup = self.get_second_cup_qualified_teams()
        second_cup_prev = self.get_second_cup_prev_qualified_teams()
        third_cup = self.get_third_cup_qualified_teams()
        third_cup_prev = self.get_third_cup_prev_qualified_teams()
        return ([first_cup, second_cup, third_cup, first_cup_prev, second_cup_prev, third_cup_prev], self.region)

    def get_first_cup_qualified_teams(self):
        return self.first_cup_qualified
    
    def get_second_cup_qualified_teams(self):
        return self.second_cup_qualified
    
    def get_third_cup_qualified_teams(self):
        return self.third_cup_qualified
    
    def get_first_cup_prev_qualified_teams(self):
        return self.first_cup_prev_qualified
    
    def get_second_cup_prev_qualified_teams(self):
        return self.second_cup_prev_qualified
    
    def get_third_cup_prev_qualified_teams(self):
        return self.third_cup_prev_qualified
    
    def merge_tables(self, tables):
        merged = defaultdict(lambda: {
            "pts": 0,
            "w": 0,
            "l": 0,
            "d": 0,
            "gf": 0,
            "gc": 0,
            "gd": 0
        })

        for table in tables:              # cada "tabla"
            for team, stats in table:     # cada tupla ("team", {...})
                for key, value in stats.items():
                    merged[team][key] += value

        # Convertir al formato original: lista de tuplas
        result = [(team, stats) for team, stats in merged.items()]
        result.sort(key=lambda x: (x[1]["pts"], x[1]["gd"], x[1]["gf"]), reverse=True)
        return result

    def generate_tournament_excel(self, file_path="tournament_simulation.xlsx"):
        '''
        merged_tables = self.merge_tables(self.general_tables)
        self.general_tables.append(merged_tables)
        self.element_names.append('Tabla General')
        self.general_matches.append([])
        '''
        wb = Workbook()
        wb.remove(wb.active)

        for i in range(len(self.element_names)):

            sheet_name = self.element_names[i][:31]  # Excel limite nombre hoja
            ws = wb.create_sheet(title=sheet_name)

            table = self.general_tables[i]
            matches = self.general_matches[i]

            # -------- TABLA --------
            ws["A1"] = "Tabla"
            ws["A1"].font = Font(bold=True)

            headers = ["Pos", "Team", "Pts", "W", "D", "L", "GF", "GC", "GD"]

            for col, header in enumerate(headers, start=1):
                ws.cell(row=2, column=col, value=header).font = Font(bold=True)

            row = 3
            pos = 1

            for team, stats in table:

                ws.cell(row=row, column=1, value=pos)
                ws.cell(row=row, column=2, value=team)
                ws.cell(row=row, column=3, value=stats["pts"])
                ws.cell(row=row, column=4, value=stats["w"])
                ws.cell(row=row, column=5, value=stats["d"])
                ws.cell(row=row, column=6, value=stats["l"])
                ws.cell(row=row, column=7, value=stats["gf"])
                ws.cell(row=row, column=8, value=stats["gc"])
                ws.cell(row=row, column=9, value=stats["gd"])

                row += 1
                pos += 1

            # -------- PARTIDOS --------
            start_row = row + 2

            ws.cell(row=start_row, column=1, value="Partidos").font = Font(bold=True)

            match_headers = ["Team 1", "Team 2", "Score 1", "Score 2"]

            for col, header in enumerate(match_headers, start=1):
                ws.cell(row=start_row + 1, column=col, value=header).font = Font(bold=True)

            r = start_row + 2
            if isinstance(matches, dict):
                matches = [matches]
            for match in matches:

                for col, key in enumerate(["team1","team2","score1","score2"], start=1):
                    ws.cell(row=r, column=col, value=match[key])

                r += 1

        wb.save(file_path)

        return file_path
    

