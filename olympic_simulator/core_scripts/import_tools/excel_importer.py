from openpyxl import load_workbook


class ExcelImporter:

    def __init__(self, archivo):
        self.workbook = load_workbook(archivo)

    def read(self):
        bloques = []

        for hoja in self.workbook.worksheets:
            years = []

            # Buscar todos los años
            for merged in hoja.merged_cells.ranges:
                fila = merged.min_row
                valor = hoja.cell(fila, merged.min_col).value
                if isinstance(valor, int):
                    years.append((valor,fila))

            years.sort(key=lambda x: x[1])
            for i, (year, fila_inicio) in enumerate(years):
                if i == len(years)-1:
                    fila_fin = hoja.max_row
                else:
                    fila_fin = years[i+1][1]-1

                encabezados = [c.value for c in hoja[fila_inicio+1]]
                datos = []

                for fila in hoja.iter_rows(min_row=fila_inicio+2, max_row=fila_fin, values_only=True):

                    if fila[0] is None:
                        continue

                    datos.append(dict(zip(encabezados,fila)))

                bloques.append({
                    "year":year,
                    "headers":encabezados,
                    "rows":datos
                })

        return bloques
    
    def ranking_prueba(self, tabla, prueba, ascendente=True):
        ranking = []
        for atleta in tabla["rows"]:
            ranking.append({
                "participant": atleta["País"],
                "value": atleta[prueba]
            })
        ranking.sort(
            key=lambda x: x["value"],
            reverse=not ascendente
        )

        return ranking[:3]